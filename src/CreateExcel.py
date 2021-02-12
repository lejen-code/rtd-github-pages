from xlwt import Workbook
import openpyxl

# Workbook is created
wb = Workbook()

# add_sheet is used to create sheet.
sheet1 = wb.add_sheet('Sheet 1')

sheet1.write(1, 0, '0302833297')
sheet1.write(2, 0, '9417306607')
sheet1.write(3, 0, '3029726547')
sheet1.write(4, 0, '0239440773')
sheet1.write(5, 0, '7950677319')
sheet1.write(0, 1, '7427375303')
sheet1.write(0, 2, '8992945223')
sheet1.write(0, 3, '5342124133')
sheet1.write(0, 4, '1231233242')
sheet1.write(0, 5, '8273738183')

wb.save('xlwt example.xls')

# Workbook is created
openpyxl_wb = openpyxl.Workbook()

# create_sheet is used to create sheet.
sheet1 = openpyxl_wb.create_sheet("Sheet 1")

sheet1.write(1, 0, '0302833297')
sheet1.write(2, 0, '9417306607')
sheet1.write(3, 0, '3029726547')
sheet1.write(4, 0, '0239440773')
sheet1.write(5, 0, '7950677319')
sheet1.write(0, 1, '7427375303')
sheet1.write(0, 2, '8992945223')
sheet1.write(0, 3, '5342124133')
sheet1.write(0, 4, '1231233242')
sheet1.write(0, 5, '8273738183')

openpyxl_wb.save("openpyxl example.xlsx")